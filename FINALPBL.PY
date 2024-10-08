import tkinter as tk
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
"""workbook = openpyxl.load_workbook('time_table.xlsx')

for sheet in workbook:
    sheet.sheet_state = 'visible'

for sheet in workbook: #loop0
    if sheet['A1'].value != "DAY": #if not a1 value = DAY them
        workbook.remove(sheet) #remove/delete

# Save the updated workbook
workbook.save('timetable_updated.xlsx') #save the updated workbook"""
workbook2 = openpyxl.load_workbook('timetable_updated.xlsx')

"""
table_list = ('Table 2',
'Table 7',
'Table 11',
'Table 15',
'Table 20',
'Table 25',
'Table 30',
'Table 35',
'Table 40',
'Table 45',
'Table 50',
'Table 54',
'Table 58',
'Table 63',
'Table 68',
'Table 73')"""

"""print(table_list[div_index])"""

"""for sheet_name in workbook2.sheetnames : #to print the sheetnames
    print(sheet_name)"""

"""sheet_A = workbook2.worksheets[0] #accessed the 1st worksheet using index
cell = sheet_A['C3']
print(cell.value)"""
"""
time_no_dict = {'8:00': '1',
                '9:00': '2',
                '10:15': '3',
                '11:15': '4',
                '1:15' : '5',
                '2:15' : '6',
                '3:30': '7',
                '4:30' : '8',
                '5:30' : '9'}
no_column_dict = {
    '1': 'B',
    '2': 'D',
    '3': 'G',
    '4': 'I',
    '5': 'L',
    '6': 'N',
    '7': 'Q',
    '8': 'S',
    '9': 'U'
}"""








#new object of tk class
main_window = tk.Tk()

# title of our window
main_window.title("PYTHON PBL")

main_window.geometry("1200x800")
bg_image = tk.PhotoImage(file="riyalestofriyals.png")
background_label = tk.Label(main_window, image=bg_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

label = tk.Label(main_window, text = "VIIT CLASSROOM TRACKER", font=("Arial", 22))
label.pack()

# label for DIV
label_DIV = tk.Label(main_window, text="DIV ", font=("Arial", 14))
label_DIV.place(x=100, y=100)

# input for DIV
DIV_input = tk.StringVar()

# entry widget for the input
DIV_entry = tk.Entry(main_window, textvariable=DIV_input, font=("Arial", 14))
DIV_entry.place(x=400, y=100)

# label for BATCH
label_BATCH = tk.Label(main_window, text="BATCH ", font=("Arial", 14))
label_BATCH.place(x=100, y=150)

# input for BATCH
BATCH_input = tk.StringVar()

# entry widget for the input
BATCH_entry = tk.Entry(main_window, textvariable=BATCH_input, font=("Arial", 14))
BATCH_entry.place(x=400, y=150)


# label for TIME
label_TIME = tk.Label(main_window, text="TIME ", font=("Arial", 14))
label_TIME.place(x=100, y=200)

# input for TIME
TIME_input = tk.StringVar()

# entry widget for the input
TIME_entry = tk.Entry(main_window, textvariable=TIME_input, font=("Arial", 14))
TIME_entry.place(x=400, y=200)



# label for DAY
label_DAY = tk.Label(main_window, text="DAY ", font=("Arial", 14))
label_DAY.place(x=100, y=250)

# input for DAY
DAY_input = tk.StringVar()

# entry widget for the input
DAY_entry = tk.Entry(main_window, textvariable=DAY_input, font=("Arial", 14))
DAY_entry.place(x=400, y=250)



time_list = ['8:00',
             '9:00',
             '10:15',
             '11:15',
             '1:15',
             '2:15',
             '3:30',
             '4:30',
             '5:30']

column_list = [
    'B',
    'D',
    'G',
    'I',
    'L',
    'N',
    'Q',
    'S',
    'U']


row_dict = {
    'MON': '3',
    'TUE': '9',
    'WED': '15',
    'THU': '21',
    'FRI': '27',
    'SAT': '33',
    'SUN': '39'
}
row_dict2 = {
    'MON': '4',
    'TUE': '10',
    'WED': '16',
    'THU': '22',
    'FRI': '28',
    'SAT': '34',
    'SUN': '40'
}
batches = ['A1',
           'A2',
           'A3',
           'B1',
           'B2', 'B3', 'C1', 'C2', 'C3', 'D1', 'D2', 'D3', 'E1', 'E2', 'E3', 
           'F1', 'F2', 'F3', 'G1', 'G2', 'G3', 'H1', 'H2', 'H3', 'I1', 'I2', 'I3',
           'J1', 'J2', 'J3', 'K1', 'K2', 'K3', 'L1', 'L2', 'L3', 'M1', 'M2', 'M3', 
           'N1', 'N2', 'N3', 'O1', 'O2', 'O3', 'P1', 'P2', 'P3', 'Q1', 'Q2', 'Q3', 
           ]
#To get the division
div_list = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P')
div = DIV_input.get()
batch = BATCH_input.get()


def timetable_teller(div, batch):
    if batch in batches:
        day = DAY_input.get()
        time = str(TIME_input.get())
        
        sheet_index = div_list.index(div)
        sheet_A = workbook2.worksheets[sheet_index]
        
        # For lecture timetable
        column_index_lecture = column_index_from_string(column_list[time_list.index(time)])
        next_column_lecture = get_column_letter(column_index_lecture + 1)
        cell_of_interest_lecture = next_column_lecture + row_dict[day]
        value_of_cell_of_interest_lecture = sheet_A[cell_of_interest_lecture].value
        print("The lecture is :",value_of_cell_of_interest_lecture)
        
        # For class timetable
        column_index = column_index_from_string(column_list[time_list.index(time)])
        next_column = get_column_letter(column_index)
        cell_of_interest = next_column + row_dict2[day]
        value_of_cell_of_interest_class = sheet_A[cell_of_interest].value
        print("The Class room is :",value_of_cell_of_interest_class)

        lecture_output_label.configure(
            text = "The lecture is of : " + value_of_cell_of_interest_lecture,
            font=("Arial", 15))

        class_output_label.configure(
            text= "The class is in : " + value_of_cell_of_interest_class,
            font=("Arial", 15))
    else:
        print("Invalid batch value.")

class_output_label = tk.Label(main_window)
class_output_label.place(x = 200, y = 600)
class_output_label.config(width=50, height=2)

lecture_output_label = tk.Label(main_window)
lecture_output_label.place(x = 200, y = 700)
lecture_output_label.config(width=50, height=2)

estimate_button = tk.Button(main_window, text="Tell!", font="Arial", command=lambda: timetable_teller(DIV_input.get(), BATCH_input.get()), width=10, height=5)
estimate_button.place(x = 900, y = 300 )
main_window.mainloop()